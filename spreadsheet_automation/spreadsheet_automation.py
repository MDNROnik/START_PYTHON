import openpyxl as xl
from openpyxl.chart import  BarChart, Reference


wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]

# show the OLD sheet data
for row in range(1, sheet.max_row+1):
    for col in range(1, sheet.max_column+1):
        print(sheet.cell(row, col).value, end=" - ")
    print()
print()

for row in range(2, sheet.max_row+1):
    value = sheet.cell(row, 3).value
    new_value = float(value * 0.9)
    new_coloumn = sheet.cell(row,4)
    new_coloumn.value = new_value
    # print(sheet.max_column, new_value)

val = Reference(sheet,
                min_row=2,
                max_row=sheet.max_row,
                min_col=4,
                max_col=4
                )

chart = BarChart()
chart.add_data(val)
sheet.add_chart(chart,"e2")

wb.save("transactions.xlsx")
